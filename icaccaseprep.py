import os
import getpass
import shutil
from datetime import datetime
from pdfminer.high_level import extract_text
from fillpdf import fillpdfs
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle


username = getpass.getuser()
now = datetime.now()
dt_string = now.strftime("%m/%d/%Y %H:%M:%S")
dt_start = now.strftime('%m/%d/%Y')

seedpath = ('') #make to subpoenas folder
downloadspath =  ('') #make path to downloads folder
esplist = []
icacexcel = '' #make path to spreadsheet
newcasestyle = NamedStyle(name = 'newcasestyle')
newcasestyle.font = Font(name= 'Carlito', color= '00008000', size = 14) #update format
newcasestyle.fill = PatternFill('solid', fgColor='00CCFFCC') #update format
newcasestyle.border = Border(bottom=Side(style= 'thin', color= '00000000'))
newcasestyle.alignment = Alignment(horizontal= 'center', vertical= 'center')
book = openpyxl.load_workbook(icacexcel)
sheet = book.worksheets[0]
try:
    book.add_named_style(newcasestyle)
except:
    pass

print('Hello, ' + username.upper() + '! Please wait while I check your Downloads folder for Cyber Tips. \n')

for espdraft in os.listdir(seedpath):
    espname = espdraft[9:-4]
    esplist.append(espname)

for cybertips in os.listdir(downloadspath):
    tippath = os.path.join(downloadspath, cybertips)
    ctname = os.path.splitext(cybertips)[0]
    if len(ctname) > 7 and len(ctname) < 11 and ctname.isnumeric() == True:
        spreadsheet = {'Detective':username, 'Date':dt_start, 'Tip Number':[], 'ESP':[]}
        print('Now handling Cyber Tip ' + ctname + ':')
        spreadsheet['Tip Number'].append(ctname)
        casepath = () #make path to case folders
        os.mkdir(casepath)   
        shutil.move(tippath, casepath)
        ctpdfpath = os.path.join(casepath + '/' + ctname + '.pdf')
        cttext = extract_text(ctpdfpath)
        for esp in esplist:
            if esp in cttext and len(esp) > 1:
                print('    Cyber Tip ' + ctname + ' may require data from ' + esp + '.')
                spreadsheet['ESP'].append(esp)
                espfile = 'Subpoena_' + esp + '.pdf'
                seedstart = os.path.join(seedpath, espfile)
                seedend = os.path.join(casepath, espfile)
                shutil.copy(seedstart, seedend)
                print('    Filling subpoena fields for date and tip number.')
                ctformfill = 'Cyber Tip ' + ctname
                data_dict = {'Text3' : ctformfill, 'DATED' : dt_start, 'Text5~' : dt_string}
                fillpdfs.write_fillable_pdf(seedend, seedend, data_dict)
        print('    Writing data to ' + icacexcel + '\n')
        sheet.insert_rows(idx=3)
        sheet['A3'] = ctname
        sheet['C3'] = username
        sheet['F3'] = spreadsheet['Date']
        sheet['H3'] = spreadsheet['ESP'][0]
        sheet['I3'] = spreadsheet['ESP'][1]
        for row in sheet.iter_rows():
            for cell in row:
                if '3' in cell.coordinate:
                    cell.style = 'newcasestyle'
        book.save(icacexcel)

print('\nAll Cyber Tips in your Downloads folder have been processed.')